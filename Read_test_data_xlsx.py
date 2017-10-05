# -*- coding: utf-8 -*-
"""
Created by Aedon Mansfield in September 2017
Modified and extended by Weigang Wei in October 2017

Revision includes: 
    * scan all the drive
    * add main function
    * save results file in specified directory
    * add condition to decide wheterh it's a sound test
    * walk only through folder "Calculaitons"
"""
import os
from openpyxl import load_workbook
import xlrd

import numpy as np

class ScanCalculationFolders():
    def __init__(self, kJobs):
        ''' kjobs = N:\1000 - 1999'''
        self.kJobs = kJobs
        self.jobPaths = self._find_a_job() # Job folders N:\4000 - 4999\4678 Bradley Hall Farm
        self.possibleTests = self._find_soundTest_jobs()
        
    def _find_a_job(self):
        hJobFolders = os.listdir(self.kJobs)
        jobPaths = []
        for hj in hJobFolders:
            jobFolders = os.listdir(os.path.join(self.kJobs,hj))
            for job in jobFolders:
                jobPaths.append(os.path.join(self.kJobs,hj,job))
        return jobPaths
                
    def _find_soundTest_jobs(self):
        ''' scan Calculations folder, if a norbuild .nbp file is found, will add
            to the list of possible tests
            return a list of [{"root":root, "saveName": fileName}, {...}]
        '''
        possibleTests = [] #list of [{"root":root, "fileName": fileName, ...}]
        for n, jp in enumerate(self.jobPaths):
            CalculationsFolder = os.path.join(jp, "Calculations")
            for root, dirs, files in os.walk(CalculationsFolder):
                for fl in files:
                    if ".nbp" in fl:
                        possibleTests.append({"root":root})
                        break
        return possibleTests

class ExtractTestsFromXlsx(ScanCalculationFolders):
    def __init__(self, kJobs):
        self.freq = np.array([100,125,160,200,250,315,400,500,630,800,1000,1250,1600,2000,2500,3150])
        ScanCalculationFolders.__init__(self,kJobs)
        
    def _find_xlsx_or_xls_files(self, path, ends):
        ''' path = the full path of the file N:\ab.xlsx or C:\cd.xls
            ends = either ".xlsx" or ".xls"
        '''
        xlsx_files = []  # All xlsx files
        xlsx_files_full = []  # All xlsx files and path
        for root, dirs, files in os.walk(path):
            for name in files:
                if name.endswith(ends):
                    xlsx_files.append(name)
                    xlsx_files_full.append(str(root)+'\\'+str(name))
        return xlsx_files_full
                    
    def test_the_file(self,file_full): 
        """
        Test xlsx file         
        Tests whether the workbook has a 'data' worksheet.
        If it does, it adds the file name to the list files2process
        """
        files2process_full_walls = []
        files2process_full_floors = []
        path, file = os.path.split(file_full[0])
        print("processing ", path)
        for ff in file_full:
            pt,fil = os.path.split(ff)
            print("   ", fil)
            wb = load_workbook(ff, read_only=True)
            try:
                ws = wb['Data']
                impactTestFound = self._exclude_impact_tests(ws['B32'].value)
                if impactTestFound==0:
                    partition = self._determine_wall_floor(ff, ws)
                    if partition == "Wall":
                        files2process_full_walls.append(ff)
                    elif partition == "Floor":
                        files2process_full_floors.append(ff)
                    else:
                        print('Partition not identified !')
            except:
                print("  Not sound test")
        return files2process_full_walls, files2process_full_floors
            
    def _test_xls_file(self, file_full):
        ''' test xls file '''
        files2process_full_walls = []
        files2process_full_floors = []
        path, file = os.path.split(file_full[0])
        print("processing ", path)
        for ff in file_full:
            print("   ", ff)
            wb = xlrd.open_workbook(ff)
            try:
                ws = wb.sheet_by_name('Data')
                impactTestFound = self._exclude_impact_tests(ws.cell(31,1).value)
                if impactTestFound==0:
                    partition = self._determine_wall_floor(ff, ws)
                    if partition == "Wall":
                        files2process_full_walls.append(ff)
                    elif partition == "Floor":
                        files2process_full_floors.append(ff)
                    else:
                        print('Partition not identified !')
            except:
                print("  Not sound test")
        return files2process_full_walls, files2process_full_floors
        
            
    def _extract_DnT_from_XLSX(self,file_full, ID):
        """Extracts the DnT values from the xlsx workbooks which have 'data' worksheet"""
        wb = load_workbook(file_full, read_only=True)
        pt, fil = os.path.split(file_full)
        Visit = fil.split(".")[0]
        ws = wb['Data']
        DnT = np.array([ws['D5'].value, ws['D6'].value, ws['D7'].value,
                        ws['D8'].value, ws['D9'].value, ws['D10'].value, 
                        ws['D11'].value, ws['D12'].value, ws['D13'].value,
                        ws['D14'].value, ws['D15'].value, ws['D16'].value,
                        ws['D17'].value, ws['D18'].value, ws['D19'].value,
                        ws['D20'].value])
        Dntw = int(ws['B39'].value)
        Ctr = int(ws['B41'].value)
        DntwCtr = Dntw + Ctr
        d = {'Visit':Visit, 'Test':ID, 'DnT': DnT, 'Dntw + Ctr': DntwCtr, 'Dntw': Dntw}
        return d
        
    def _extract_DnT_from_XLS(self,file_full, ID):
        """Extracts the DnT values from the xls workbooks which have 'data' worksheet"""
        wb = xlrd.open_workbook(file_full)
        pt, fil = os.path.split(file_full)
        Visit = fil.split(".")[0]

        ws = wb.sheet_by_name('Data')
        col = 3
        rows = [4,5,6,7,8,9,10,11,12,13,14,15,16,17,18,19]
        DnT = np.array([ws.cell(r,col).value for r in rows])
        Dntw = int(ws.cell(38,1).value)
        Ctr = int(ws.cell(40,1).value)
        DntwCtr = Dntw + Ctr
        d = {'Visit':Visit, 'Test':ID, 'DnT': DnT, 'Dntw + Ctr': DntwCtr, 'Dntw': Dntw}
        return d
    
    def _exclude_impact_tests(self,B32C31):
        ''' B32C31 = String 
                        the value of B32 if xlsx file, and
                        the value of cell(31,1) if xls file
        '''
        if '140-7' in B32C31:
            print('Impact test')
            impactTestFound = 1
        else:
            impactTestFound = 0            
        return impactTestFound
                
    def _determine_wall_floor(self, xlsxFile, ws):
        if xlsxFile.endswith('.xlsx'):
            if 'Floor' in ws['B36'].value or 'floor' in ws['B36'].value:
                return "Floor"
            elif 'Wall' in ws['B36'].value or 'wall' in ws['B36'].value:
                return "Wall"
            else:
                return "Unknown"
        elif xlsxFile.endswith('.xls'):
            if 'Floor' in ws.cell(35,1).value or 'floor' in ws.cell(35,1).value:
                return "Floor"
            elif 'Wall' in ws.cell(35,1).value or 'wall' in ws.cell(35,1).value:
                return "Wall"
            else:
                return "Unknown"   
    
    def _write_data_to_CSV(self,files2process_full, pKey):
        ds = []
        for i,fpath in enumerate(files2process_full):
            p, fil = os.path.split(fpath)
            print(i, '  geting data from ', fil)
            if fpath.endswith(".xlsx"):
                ds.append(self._extract_DnT_from_XLSX(fpath, i))
            else:
                ds.append(self._extract_DnT_from_XLS(fpath, i))
                
        print("\nGot data. Start to writing...\n")
        path,folderName = os.path.split(files2process_full[0])
        pathStringList = path.split("\\")
        saveFile = ''
        for m in range(len(pathStringList)-1):
            saveFile = saveFile + pathStringList[m+1] +'-'
        saveFile = saveFile + pKey + '.CSV'
        
        with open(saveFile, 'w') as ff:
            ff.write('{}, '.format('Frequency'))
            for d1 in ds:
                ff.write('{} - Test {}, '.format(d1['Visit'], d1['Test']))
            for f, fr in enumerate(self.freq):
                ff.write("\n")
                ff.write('{}, '.format(fr))
                for d1 in ds:                
                    ff.write('{}, '.format(d1['DnT'][f]))
            ff.write("\n\n")
            ff.write('{}, '.format('Dntw'))
            for d1 in ds:
                ff.write('{}, '.format(d1['Dntw']))
            ff.write("\n")
            ff.write('DnTw+Ctr, ')
            for d1 in ds:
                ff.write('{}, '.format(d1['Dntw + Ctr']))
            ff.write("\n")  
        print ("\nDone!\n")
            
    def _run_the_scan(self):
        for pt in self.possibleTests:
            root = pt['root']
            xlsx_files_full = self._find_xlsx_or_xls_files(root, ".xlsx")
            xls_files_full = self._find_xlsx_or_xls_files(root, ".xls")
            
            if len(xlsx_files_full)>0:
                print('xlsx found!')
                wallFiles, floorFiles = self.test_the_file(xlsx_files_full)
                if (len(wallFiles)>0):
                    self._write_data_to_CSV(wallFiles, "Wall")
                if (len(floorFiles)>0):
                    self._write_data_to_CSV(floorFiles, "Floor")
            if len(xls_files_full)>0:
                print('xls found!')
                wallFiles, floorFiles  = self._test_xls_file(xls_files_full)
                if (len(wallFiles)>0):
                    self._write_data_to_CSV(wallFiles, "Wall")
                if (len(floorFiles)>0):
                    self._write_data_to_CSV(floorFiles, "Floor")
                
            
def run_scan():
    ex = ExtractTestsFromXlsx(r"N:\Staff\Weigang\scripts\extract-measurements\4000 - 4999")
    ex._run_the_scan()
    
    
if __name__=="__main__":
    run_scan()
    